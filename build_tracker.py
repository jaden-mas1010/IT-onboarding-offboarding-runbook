import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- Sheet 1: SLA Tracker ----------
ws1 = wb.active
ws1.title = "SLA Tracker"

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
title_font = Font(name="Arial", bold=True, size=14)
normal_font = Font(name="Arial", size=10)
thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

ws1["A1"] = "IT Ticket SLA Tracker"
ws1["A1"].font = title_font
ws1.merge_cells("A1:I1")

headers = ["Ticket ID", "Type", "Priority", "Opened", "SLA Target (hrs)", "Resolved", "Resolution Time (hrs)", "Within SLA?", "Notes"]
for col, h in enumerate(headers, start=1):
    c = ws1.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", wrap_text=True)

data = [
    ["IT-1001", "Onboarding - Device Setup", "Medium", "2026-08-19 09:14", 72, "2026-08-19 11:45", None, None, "New hire Windows laptop"],
    ["IT-1002", "Onboarding - M365 Provisioning", "Medium", "2026-08-19 12:00", 72, "2026-08-19 13:00", None, None, "New hire M365 account"],
    ["IT-1003", "Offboarding", "High", "2026-08-21 08:30", 8, "2026-08-21 16:50", None, None, "Planned departure, full lifecycle"],
    ["IT-1004", "Access Issue", "Medium", "2026-08-20 10:05", 4, "2026-08-20 10:45", None, None, "Shared drive permission gap"],
]

from datetime import datetime
for i, row in enumerate(data, start=4):
    for j, val in enumerate(row, start=1):
        cell = ws1.cell(row=i, column=j, value=val)
        cell.font = normal_font
        cell.border = thin_border
    # Resolution Time (hrs) = (Resolved - Opened) * 24
    ws1.cell(row=i, column=7, value=f"=(F{i}-D{i})*24")
    ws1.cell(row=i, column=7).number_format = "0.00"
    ws1.cell(row=i, column=7).font = normal_font
    # Within SLA? = IF(resolution_time <= sla_target, "Yes", "No")
    ws1.cell(row=i, column=8, value=f'=IF(G{i}<=E{i},"Yes","No")')
    ws1.cell(row=i, column=8).font = normal_font

# Format date columns as datetime
for row in range(4, 4+len(data)):
    for col in [4, 6]:
        cell = ws1.cell(row=row, column=col)
        cell.number_format = "yyyy-mm-dd hh:mm"

# Summary
ws1["A9"] = "Summary"
ws1["A9"].font = Font(name="Arial", bold=True, size=12)
ws1["A10"] = "Total tickets:"
ws1["B10"] = "=COUNTA(A4:A7)"
ws1["A11"] = "Tickets within SLA:"
ws1["B11"] = '=COUNTIF(H4:H7,"Yes")'
ws1["A12"] = "SLA compliance rate:"
ws1["B12"] = "=B11/B10"
ws1["B12"].number_format = "0.0%"
for r in [10, 11, 12]:
    ws1.cell(row=r, column=1).font = normal_font
    ws1.cell(row=r, column=2).font = normal_font

col_widths = [12, 26, 10, 18, 14, 18, 18, 12, 32]
for i, w in enumerate(col_widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ---------- Sheet 2: Asset Inventory ----------
ws2 = wb.create_sheet("Asset Inventory")
ws2["A1"] = "IT Asset Inventory"
ws2["A1"].font = title_font
ws2.merge_cells("A1:H1")

headers2 = ["Asset ID", "Type", "OS", "Assigned To", "Department", "Status", "Purchase Date", "Warranty Expiry"]
for col, h in enumerate(headers2, start=1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", wrap_text=True)

assets = [
    ["WIN-0142", "Laptop", "Windows 11", "New Hire - CS", "Customer Success", "Assigned", "2026-08-19", "2029-08-19"],
    ["WIN-0098", "Laptop", "Windows 11", "Unassigned", "Pool", "Available", "2025-11-02", "2028-11-02"],
    ["MAC-0031", "Laptop", "macOS Sonoma", "Unassigned", "Pool", "Returned/Wiped", "2025-06-14", "2028-06-14"],
    ["WIN-0071", "Laptop", "Windows 11", "Departed Employee", "Finance", "Returned/Wiped", "2024-09-10", "2027-09-10"],
]

for i, row in enumerate(assets, start=4):
    for j, val in enumerate(row, start=1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.font = normal_font
        cell.border = thin_border

for row in range(4, 4+len(assets)):
    for col in [7, 8]:
        ws2.cell(row=row, column=col).number_format = "yyyy-mm-dd"

col_widths2 = [12, 10, 14, 20, 18, 16, 14, 16]
for i, w in enumerate(col_widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save("assets/sla-and-asset-tracker.xlsx")
print("saved")
